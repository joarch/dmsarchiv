# -*- coding: utf-8 -*-

import json
import os
import re
import sys
from datetime import datetime, date
from decimal import Decimal
from getopt import getopt, GetoptError

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from common import _json_load


def export_nach_excel(documents, export_profil):
    # Zieldateiname ermitteln
    ziel_dateiname = export_profil["dateiname"]
    postfix = export_profil.get("dateiname_postfix")
    if postfix:
        if "%" in postfix:
            postfix = datetime.now().strftime(postfix)
        splitext = os.path.splitext(os.path.basename(export_profil["dateiname"]))
        ziel_dateiname = os.path.join(
            os.path.dirname(export_profil["dateiname"]),
            splitext[0] + postfix + splitext[1]
        )

    # letzte fortlaufende Nummer ermitteln
    fortlaufendes_feld = export_profil.get("fortlaufendes_feld")
    letzte_fortlaufende_nummer = -1
    filename_fortlaufendes_feld = None
    if fortlaufendes_feld:
        filename_fortlaufendes_feld = os.path.join(
            os.path.dirname(export_profil["dateiname"]),
            os.path.splitext(os.path.basename(export_profil["dateiname"]))[0] + "_" +
            "fortlaufendes_feld.txt"
        )
        if os.path.exists(filename_fortlaufendes_feld):
            with open(filename_fortlaufendes_feld, 'r', encoding='utf-8') as outfile:
                value = outfile.read()
                if value:
                    letzte_fortlaufende_nummer = int(value)

    # Datei Mapping Caches
    datei_mappings = dict()

    # Zeilen und Spalten aus den Dokumenten anhand Export Profil ermitteln
    rows = list()
    for document in documents["documents"]:
        columns = list()
        rows.append(columns)

        for spalte in export_profil["spalten"]:
            column = dict()
            columns.append(column)

            feld_name = spalte["feld"]
            if spalte.get("alias"):
                column["feld_name"] = spalte["alias"]
            else:
                column["feld_name"] = spalte["feld"]
            mapped_value = ""

            if feld_name:
                # Spalten Wert auslesen und mappen
                if feld_name in document:
                    value = document[feld_name]
                elif feld_name in document["classifyAttributes"]:
                    value = document["classifyAttributes"][feld_name]
                else:
                    raise RuntimeError(
                        f"Die Spalte '{feld_name}' existiert nicht im Dokument. Bitte Export-Profil überprüfen.")

                # Mapping
                mapping_def = spalte.get("mapping")
                if mapping_def is not None:
                    # konfiguriertes Mapping anwenden
                    if mapping_def["typ"] == "re":
                        # Mapping mit RegEx Methode
                        # - zuerst immer in String umwandeln, das Mapping geht aktuell nur mir RegEx
                        mapped_value = map_value(value, "string")
                        re_operation = getattr(re, mapping_def["methode"])
                        argumente = mapping_def["argumente"]
                        if len(argumente) == 2:
                            mapped_value = re_operation(argumente[0], argumente[1], mapped_value)
                        else:
                            raise RuntimeError(
                                f"Fehler beim Mapping zum Feld '{feld_name}'. "
                                f"Es werden nur 2 Argument unterstützt.")
                        mapped_value = map_value(mapped_value, spalte.get("type"))
                    elif mapping_def["typ"] == "datei":
                        # Mapping aus Datei auslesen
                        # - Datei Mapping Cache initialisieren
                        if datei_mappings.get(mapping_def["dateiname"]) is None:
                            datei_mappings[mapping_def["dateiname"]] = _init_mapping_data(mapping_def)
                        # mapping von id zu name
                        mapping_data = datei_mappings[mapping_def["dateiname"]]
                        mapped_value = mapping_data[value]
                    else:
                        raise RuntimeError(f"Unbekannter Mapping Typ: {mapping_def['type']}")
                else:
                    mapped_value = map_value(value, spalte.get("type"))
            else:
                # keine Feld Name, damit bleibt die Spalte leer
                pass

            column["value"] = mapped_value
            if spalte.get("number_format"):
                column["number_format"] = spalte["number_format"]
            else:
                if isinstance(mapped_value, date):
                    column["number_format"] = 'DD.MM.YYYY'
                if isinstance(mapped_value, datetime):
                    column["number_format"] = 'DD.MM.YYYY HH:MM:SS'
            if spalte.get("computed"):
                column["computed"] = spalte["computed"]

    # sortieren
    if export_profil.get("sortierung"):
        for sort_def in reversed(export_profil["sortierung"]["felder"]):
            if sort_def["wie"] == "absteigend":
                reverse = True
            elif sort_def["wie"] == "aufsteigend":
                reverse = False
            else:
                raise RuntimeError(
                    f"Unbekannte Sortierung zum 'feld'='{sort_def['feld']}' mit 'wie'='{sort_def['wie']}' "
                    f", erlaubt sind nur 'aufsteigend' oder 'absteigend'.")
            rows.sort(
                key=lambda r: list(filter(lambda c: c["feld_name"] == sort_def["feld"], r))[0]["value"],
                reverse=reverse
            )

    # Computed und Format ermitteln
    for row in rows:
        for column in row:
            # computed Wert ermitteln
            if column.get("computed"):
                computed = column.get("computed")
                # bekannte Methoden ersetzen
                computed = computed \
                    .replace("nicht_fortlaufend()",
                             "pruefe_is_nicht_fortlaufend(row, fortlaufendes_feld, letzte_fortlaufende_nummer)")
                column["value"] = eval(computed)
            # Format ermitteln
            if export_profil.get("formate"):
                for format_candidate in export_profil["formate"]:
                    if re.match(format_candidate["match"], str(column["value"])):
                        if "PatternFill" == format_candidate["format"]["format"]:
                            column["fill"] = PatternFill(start_color=format_candidate["format"]["start_color"],
                                                         end_color=format_candidate["format"]["end_color"],
                                                         fill_type=format_candidate["format"]["fill_type"])

        for column in row:
            # max. fortlaufendes Feld merken
            if fortlaufendes_feld and column["feld_name"] == fortlaufendes_feld:
                letzte_fortlaufende_nummer = column["value"]
        if not letzte_fortlaufende_nummer:
            raise RuntimeError("Die fortlaufende Nummer konnte nicht ermittelt werden")

    # als Excel speichern
    if not os.path.exists(ziel_dateiname):
        # neue Excel Datei
        if not export_profil.get("vorlage_dateiname"):
            # neu
            wb = Workbook()
            ws = wb.active
        else:
            # aus Vorlage
            wb = load_workbook(filename=export_profil["vorlage_dateiname"])
            if not export_profil.get("vorlage_sheet_name"):
                ws = wb.active
            else:
                ws = wb[export_profil["vorlage_sheet_name"]]

        row_idx = 1
        # mit Spaltenüberschrifen
        if export_profil["spaltenueberschrift"].lower() == "ja":
            column_header_format = export_profil.get("spaltenueberschrift_format")
            if column_header_format is not None:
                if "PatternFill" == column_header_format["format"]:
                    column_header = PatternFill(start_color=column_header_format["start_color"],
                                                end_color=column_header_format["end_color"],
                                                fill_type=column_header_format["fill_type"])
                else:
                    raise RuntimeError(
                        f"Unbekanntes Format {column_header_format['format']} in 'spaltenueberschrift_format/format'. "
                        f"Möglich ist nur 'PatternFill'")
            else:
                # Standard Format
                column_header = PatternFill(start_color='AAAAAA',
                                            end_color='AAAAAA',
                                            fill_type='solid')
            column_idx = 1
            for spalte in export_profil["spalten"]:
                ws.cell(column=column_idx, row=row_idx, value=spalte["ueberschrift"])
                col = ws["{}{}".format(get_column_letter(column_idx), row_idx)]
                col.font = Font(bold=True)
                col.fill = column_header
                column_idx += 1
            row_idx += 1

        # Zeilen und Spalten ins Excel Dokument schreiben
        append_rows(row_idx, rows, ws)
    else:
        # vorhandene Excel Datei fortschreiben
        wb = load_workbook(filename=ziel_dateiname)
        if not export_profil.get("vorlage_sheet_name"):
            ws = wb.active
        else:
            ws = wb[export_profil["vorlage_sheet_name"]]
        id_feld = export_profil["id_feld"]
        id_feld_idx = -1
        for idx, spalte in enumerate(export_profil["spalten"]):
            if spalte["feld"] == id_feld:
                id_feld_idx = idx
        if id_feld_idx == -1:
            raise RuntimeError(
                f"Fehler das id_feld '{id_feld}' existiert nicht als Spalte in der Export Konfiguration.")
        # update Rows
        empties = 0
        last_row = 0
        for row_idx, row in enumerate(ws.iter_rows()):
            cell = row[id_feld_idx]
            if cell.value:
                empties = 0
                update_row(cell.value, id_feld_idx, rows, row)
                rows = remove_row(cell.value, id_feld_idx, rows)
            else:
                empties += 1
                for cell in row:
                    # evtl. leere Zeile, nur wenn alle Spalten ebenfalls leer sind
                    if cell.value:
                        empties = 0
                        break
            if empties == 0:
                last_row = row_idx + 1
            if empties > 100:
                # fertig, nur noch leere Id Spalten
                break
        # neue Rows anhängen
        row_idx = last_row + 1
        append_rows(row_idx, rows, ws)

    wb.save(filename=ziel_dateiname)
    print(f"Die Excel-Datei wurde geschrieben: '{ziel_dateiname}'")

    # letzte fortlaufende Nummer in Datei merken
    if fortlaufendes_feld:
        with open(filename_fortlaufendes_feld, 'w', encoding='utf-8') as outfile:
            outfile.write(str(letzte_fortlaufende_nummer))


def append_rows(row_idx, rows, ws):
    """
    Hängt die rows an das Sheet, beginnend ab Zeile row_idx
    """
    for row in rows:
        column_idx = 1
        for column in row:
            new_cell = ws.cell(column=column_idx, row=row_idx, value=column["value"])
            if column.get("number_format"):
                new_cell.number_format = column["number_format"]
            if column.get("fill"):
                new_cell.fill = column["fill"]
            column_idx += 1
        row_idx += 1


def update_row(id_value, id_feld_idx, rows, row):
    """
    Aktualisiert die row im Sheet, wenn sie innerhalb der neuen rows existiert.
    Existiert die row nicht in den neuen rows, bleibt sie unverändert.
    """
    existing = list(filter(lambda r: r[id_feld_idx]["value"] == id_value, rows))
    if len(existing) == 1:
        # aktualisieren vorhandene Row
        for column_idx, column in enumerate(existing[0]):
            row[column_idx].value = column["value"]
            if column.get("number_format"):
                row[column_idx].number_format = column["number_format"]
            if column.get("fill"):
                row[column_idx].fill = column["fill"]
    elif len(existing) > 1:
        raise RuntimeError(f"Zeile mit Id '{id_value}' ist mehrfach vorhanden. Anzahl: {len(existing)}")
    # ignorieren, Row nur im Excel Dokument


def remove_row(id_value, id_feld_idx, rows):
    """
    Löscht die row mit der id_value aus den rows
    """
    return [row for row in rows if row[id_feld_idx]["value"] != id_value]


def map_value(value, mapping_type=None):
    if mapping_type == "string":
        return str(value)
    if mapping_type == "int":
        try:
            return int(value)
        except ValueError:
            return -1

    return map_str_value(value)


def map_str_value(value):
    if type(value) != str:
        return value

    if value == "undefined":
        # clean up
        value = ""

    if value == "true":
        value = "ja"

    if value == "false":
        value = "nein"

    if "€" in value \
            and (value[0].isnumeric() or len(value) >= 2 and value[0] == "-" and value[1].isnumeric()):
        return map_eur(value)

    eur_pattern = re.compile(r"^-?[0-9]+,?[0-9]* (€|EUR)$")
    if eur_pattern.match(value):
        return map_eur(value)

    datum_pattern = re.compile(r"^[0-9]{2}\.[0-9]{2}\.[0-9]{4}$")
    if datum_pattern.match(value):
        return map_datum(value)

    datum_pattern = re.compile(r"^[0-9]{4}-[0-9]{2}-[0-9]{2}$")
    if datum_pattern.match(value):
        return map_datum(value)

    datum_pattern = re.compile(r"^[0-9]{2}\.[0-9]{2}\.[0-9]{4} [0-9]{2}:[0-9]{2}:[0-9]{2}$")
    if datum_pattern.match(value):
        return map_datum_zeit(value)

    datum_pattern = re.compile(r"^[0-9]{4}-[0-9]{2}-[0-9]{2} [0-9]{2}:[0-9]{2}:[0-9]{2}$")
    if datum_pattern.match(value):
        return map_datum_zeit(value)

    decimal_pattern = re.compile(r"^-?[0-9]+,?[0-9]*$")
    if decimal_pattern.match(value):
        return map_number(value)

    return value


def map_number(value):
    if value is None:
        return None
    return Decimal(value.replace('.', '').replace(' ', '').replace(',', '.'))


def map_eur(value):
    return map_number(value.replace("€", "").replace("EUR", ""))


def map_datum(value):
    if "-" in value:
        return datetime.strptime(value, "%Y-%m-%d").date()
    return datetime.strptime(value, "%d.%m.%Y").date()


def map_datum_zeit(value):
    if "-" in value:
        return datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
    return datetime.strptime(value, "%d.%m.%Y %H:%M:%S")


def pruefe_is_nicht_fortlaufend(columns, fortlaufendes_feld, previous_fortlaufendes_feld):
    return not list(filter(lambda c: c["feld_name"] == fortlaufendes_feld, columns))[0][
                   "value"] == previous_fortlaufendes_feld + 1


def _init_mapping_data(mapping_def):
    """
    List alle Einträge und erzeugt ein neues Dict anhand der 'id' und 'name' Definition
    """
    result = dict()
    mapping_data = _json_load(mapping_def["dateiname"])
    for entry in mapping_data:
        result[entry[mapping_def["id"]]] = entry[mapping_def["name"]]
    return result


def main(argv):
    """
    Export die übergebene JSON Datei (documents_datei) mit den exportierten DMS Dokumenten Feldern nach Excel.
    Das Export Format wird mit der übergebenen Export Parameter Datei (export_parameter_datei) konfiguriert.
    """
    hilfe = f"{os.path.basename(__file__)} -d <documents_datei> -e <export_parameter_datei>"
    documents_datei = ""
    export_parameter_datei = ""
    try:
        opts, args = getopt(argv, "hd:e:", ["documents_datei=", "export_parameter_datei="])
    except GetoptError:
        print(hilfe)
        sys.exit(2)

    for opt, arg in opts:
        if opt in ("-h", "--help"):
            print(hilfe)
            sys.exit()
        elif opt in ("-d", "--documents_datei"):
            documents_datei = arg
        elif opt in ("-e", "--export_parameter_datei"):
            export_parameter_datei = arg

    if not documents_datei or not export_parameter_datei:
        print("Usage: " + hilfe)
        sys.exit(2)

    if not os.path.exists(documents_datei):
        raise RuntimeError(f"Die Datei '{documents_datei}' existiert nicht.")
    if not os.path.exists(export_parameter_datei):
        raise RuntimeError(f"Die Datei '{export_parameter_datei}' existiert nicht.")

    with open(documents_datei, encoding="utf-8") as file:
        documents = json.load(file)
    with open(export_parameter_datei, encoding="utf-8") as file:
        export_parameter = json.load(file)

    export_nach_excel(documents, export_parameter["export"])


if __name__ == '__main__':
    # main(sys.argv[1:])
    main(["-d", "../export_documents.json", "-e", "../config/dmsarchiv_vorlage.json"])
